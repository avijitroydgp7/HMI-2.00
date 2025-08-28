from __future__ import annotations

from typing import Dict, Any, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from services.comment_data_service import comment_data_service


class ExcelService:
    """Import and export comment tables to Excel files."""

    # --- Reading ---------------------------------------------------------
    def read_comments_from_sheet(self, ws) -> Dict[str, Any]:
        columns = [cell.value or "" for cell in ws[1]] if ws.max_row else []
        comments: List[List[Dict[str, Any]]] = []
        for row in ws.iter_rows(min_row=2, max_col=len(columns)):
            row_data: List[Dict[str, Any]] = []
            for cell in row:
                raw = "" if cell.value is None else str(cell.value)
                fmt: Dict[str, Any] = {}
                if cell.font:
                    if cell.font.bold:
                        fmt["bold"] = True
                    if cell.font.italic:
                        fmt["italic"] = True
                    if cell.font.underline:
                        fmt["underline"] = True
                if cell.fill and getattr(cell.fill, "fgColor", None) and cell.fill.fill_type == "solid":
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != "00000000":
                        fmt["bg_color"] = f"#{rgb[-6:]}"
                # Alignment and border styles are ignored
                row_data.append({"raw": raw, "format": fmt})
            comments.append(row_data)
        return {
            "columns": columns,
            "comments": comments,
            "excel": {"sheet_name": ws.title},
        }

    def read_comments_from_file(self, file_path: str) -> Dict[str, Any]:
        wb = load_workbook(file_path)
        ws = wb.active
        return self.read_comments_from_sheet(ws)

    # --- Writing ---------------------------------------------------------
    def write_comments_to_file(self, group_id: str, file_path: str) -> bool:
        group = comment_data_service.get_group(group_id)
        if not group:
            return False
        wb = Workbook()
        ws = wb.active
        sheet_name = group.get("excel", {}).get("sheet_name", "Sheet1")
        ws.title = sheet_name
        columns = group.get("columns", [])
        # ensure latest column names are stored in the data service
        comment_data_service.update_comments(group_id, group.get("comments", []), columns)
        for c, header in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=header)
        comments = group.get("comments", [])
        for r, row in enumerate(comments, start=2):
            for c, cell_data in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=cell_data.get("raw", ""))
                fmt = cell_data.get("format", {})
                if fmt:
                    cell.font = Font(
                        bold=fmt.get("bold", False),
                        italic=fmt.get("italic", False),
                        underline="single" if fmt.get("underline") else None,
                    )
                    if "bg_color" in fmt:
                        color = fmt["bg_color"].lstrip("#")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    # Alignment and border styles are not exported
        wb.save(file_path)
        return True


excel_service = ExcelService()